from typing import Tuple, List

import openpyxl
from celery.result import AsyncResult
from django.conf import settings
from django.utils.translation import gettext_lazy as _
from openpyxl.worksheet.worksheet import Worksheet

from services.business_logic.session_models import SessionDebtorModel
from services.models import Service
from .. import utils


class FileReader:
    """Загрузка, проверка данных входящих файлов"""

    def __init__(self, service: str, task_file_verification_id: str, filename: str):
        self.service = service
        self.filename = filename
        self.task_file_verify = AsyncResult(id=task_file_verification_id)
        self.path_to_read_file = f'{settings.MEDIA_ROOT}/{self.service}/{filename}'
        self.path_to_file_with_bad_results = utils.get_results_file_abspath(service=self.service, filename=f'{self.service}:incorrect_data_or_duplicates:{filename.rsplit(":", maxsplit=1)[-1]}')

    def __call__(self, progress) -> Tuple[List[str], int]:
        """Проверяет данные в файле self.path_to_read_file, возвращает кортеж с SessionDebtorModel, множество
        уникальных значений серии номера паспорта и число дубликатов или невалидных строк"""
        incorrect_data_or_duplicates = 0
        unique_pass = set()
        bad_rows = []
        start_row = self.task_file_verify.result['ser_num_pass']['start_row']

        sheet: Worksheet = openpyxl.load_workbook(self.path_to_read_file).active
        bad_rows.append(sheet.iter_rows(min_row=1, max_row=start_row - 1, values_only=True))

        id_credit_column = self.task_file_verify.result['id_credit']['col']
        fullname_column = self.task_file_verify.result['fullname']['col']
        date_birth_column = self.task_file_verify.result['date_birth']['col']
        ser_num_pass_column = self.task_file_verify.result['ser_num_pass']['col']
        date_issue_pass_column = self.task_file_verify.result['date_issue_pass']['col']
        name_org_pass_column = self.task_file_verify.result['name_org_pass']['col']

        service_key = Service.objects.get(title=self.service).key

        for num_row in range(start_row, sheet.max_row + 1):
            fullname = str(sheet.cell(row=num_row, column=fullname_column).value) if fullname_column else ''
            id_credit = str(sheet.cell(row=num_row, column=id_credit_column).value) if id_credit_column else ''
            ser_num_pass = str(sheet.cell(row=num_row, column=ser_num_pass_column).value).replace(
                ' ', '') if ser_num_pass_column else ''

            progress.set_progress(
                current=num_row,
                total=sheet.max_row - start_row,
                description=_('Reading row') + f': {num_row}. {fullname} | {_("id credit")}: {id_credit} | '
                                               f'{_("passport")}: {ser_num_pass}'
            )
            if fullname:
                fullname = fullname.split(' ', maxsplit=2)
                if len(fullname) == 3:
                    surname, name, patronymic = fullname
                elif len(fullname) == 2:
                    surname, name, patronymic = fullname[0], fullname[1], ''
                else:
                    # Если в ячейке ФИО только одно слово
                    bad_rows.append(sheet.iter_rows(min_row=num_row, max_row=num_row, values_only=True))
                    incorrect_data_or_duplicates += 1
                    continue
                surname = surname.lower().title() if isinstance(surname, str) else surname
                name = name.lower().title() if isinstance(name, str) else name
                patronymic = patronymic.lower().title() if isinstance(patronymic, str) else patronymic
            else:
                # Если ячейка ФИО пуста или не найдена
                surname, name, patronymic = '', '', ''
                bad_rows.append(sheet.iter_rows(min_row=num_row, max_row=num_row, values_only=True))
                incorrect_data_or_duplicates += 1
                continue

            date_birth = utils.date_identifier_and_converter(
                value=sheet.cell(row=num_row, column=date_birth_column).value)[0] if date_birth_column else ''

            date_issue_pass = utils.date_identifier_and_converter(
                value=sheet.cell(row=num_row, column=date_issue_pass_column).value)[0] if date_issue_pass_column else ''

            name_org_pass = str(sheet.cell(
                row=num_row, column=name_org_pass_column).value) if name_org_pass_column else ''

            if ser_num_pass.isdigit() and len(ser_num_pass) == 10 and ser_num_pass not in unique_pass:
                unique_pass.add(ser_num_pass)
            else:
                # Дублирующийся паспорт или количество цифр не равно 10
                bad_rows.append(sheet.iter_rows(min_row=num_row, max_row=num_row, values_only=True))
                incorrect_data_or_duplicates += 1
                continue

            # Логика ниже выполнится только если есть имя и фамилия, дата рождения, серия и номер паспорта,
            # и паспорт уникальный.
            # То есть unique_pass == количеству SessionDebtorModel в
            # global_storage.storage[service][task_file_verification_id]
            if name and surname and date_birth and ser_num_pass:
                session_debtor = SessionDebtorModel(
                    surname=surname,
                    name=name,
                    patronymic=patronymic,
                    date_birth=date_birth,
                    ser_num_pass=ser_num_pass,
                    date_issue_pass=date_issue_pass,
                    name_org_pass=name_org_pass,
                    service=self.service,
                    task_file_verification_id=self.task_file_verify.task_id,
                    service_key=service_key,
                    id_credit=id_credit
                )
                if self.service == "FSSP":
                    if inn_column := self.task_file_verify.result['inn']['col']:
                        if inn := str(sheet.cell(row=num_row, column=inn_column).value):
                            session_debtor.inn = inn

            else:
                # Отсутствует значение одного из или нескольких обязательных полей
                bad_rows.append(sheet.iter_rows(min_row=num_row, max_row=num_row, values_only=True))
                incorrect_data_or_duplicates += 1

        if incorrect_data_or_duplicates > 0:
            self.bad_rows_to_file(bad_rows=bad_rows)

        return list(unique_pass), incorrect_data_or_duplicates

    def bad_rows_to_file(self, bad_rows: List) -> None:
        """Сохраняет невалидные данные в файл"""
        wb = openpyxl.Workbook()
        ws = wb.active
        for it_row in bad_rows:
            for row in it_row:
                ws.append(row)
        wb.save(self.path_to_file_with_bad_results)
