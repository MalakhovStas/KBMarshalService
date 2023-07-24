from typing import Iterable, Generator

from django.core.files.storage import FileSystemStorage
from django.core.files.uploadedfile import InMemoryUploadedFile, TemporaryUploadedFile
from celery.result import AsyncResult
from web_service.settings import redis_cache
from celery import shared_task
from celery_progress.backend import ProgressRecorder
from datetime import datetime, timedelta

from django.core.files.storage import FileSystemStorage
from django.http import HttpRequest
from django.utils.translation import gettext_lazy as _
from django.contrib import messages
from django.core.files.uploadedfile import InMemoryUploadedFile, TemporaryUploadedFile
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from django.utils import translation
from services.utils import get_redis_key


BAD_result_FNS_file = 'media/services/FNS/fns_BAD_result.xlsx'


class FileVerificationException(ValueError):
    def __init__(self, message: str = 'File verification - FAILED'):
        self.message = message
        super().__init__(self.message)


# class SessionPerson:
#     """ Класс для валидации данных должника до его записи в БД,
#     обязательные поля: name, surname, date_birth, ser_num_pass, date_issue_pass"""
#
#     def __init__(self, **kwargs):
#         self.name: str | None = kwargs.get('name')
#         self.surname: str | None = kwargs.get('surname')
#         self.patronymic: str | None = kwargs.get('patronymic')
#         self.date_birth: str | None = kwargs.get('date_birth')
#         self.ser_num_pass: str | None = kwargs.get('ser_num_pass')
#         self.date_issue_pass: str | None = kwargs.get('date_issue_pass')
#         self.name_org_pass: str | None = kwargs.get('name_org_pass')
#         self.INN: str | None = None
#         self.fns_key: str | None = kwargs.get('fns_key')
#         self.ready_for_req: bool = True if (
#                 self.surname and self.name and self.date_birth and self.ser_num_pass and self.date_issue_pass) else False
#         self.fns_url: str = f'https://api-fns.ru/api/innfl?fam={self.surname}&nam={self.name}&otch=' \
#                             f'{self.patronymic if self.patronymic else None}&bdate={self.date_birth}&' \
#                             f'doctype=21&docno={self.ser_num_pass}&key={self.fns_key}'


class BaseField:
    words_to_search_in_title = []
    or_words_to_search_in_title = []

    def get_data(self, cell_value: str):  # -> tuple[bool | str | None]:
        pass

    def check_title(self, cell_value: str, column):  # -> bool | dict:
        result = False
        cell_value = cell_value.lower()
        if all([word in cell_value for word in self.words_to_search_in_title]) or \
                all([word in cell_value for word in self.or_words_to_search_in_title]):
            result = {'column': column}
        return result

    def check_data(self, cell_value: str, row):  # -> bool | dict:
        result = False
        if any(self.get_data(cell_value=cell_value)):
            result = {'row': row}
        return result


class FullNamePerson(BaseField):
    words_to_search_in_title = ['фамилия', 'имя', 'отчество']
    or_words_to_search_in_title = ['фио']

    def get_data(self, cell_value: str) -> tuple:
        surname, name, patronymic = None, None, None
        if cell_value:
            cell_value = cell_value.split(' ', maxsplit=2)
            if 1 < len(cell_value) <= 3 and all([word.isalpha() for word in cell_value]):
                if len(cell_value) == 3:
                    surname, name, patronymic = cell_value
                elif len(cell_value) == 2:
                    surname, name, patronymic = cell_value[0], cell_value[1], ''
        return surname, name, patronymic


class DateBirthPerson(BaseField):
    words_to_search_in_title = ['дата', 'рождения']
    min_age_person = 18

    def get_data(self, cell_value: str) -> tuple:
        result = False,
        if cell_value:
            try:
                date_birth = datetime.strptime(cell_value, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    date_birth = datetime.strptime(cell_value, '%d.%m.%Y')
                except ValueError:
                    date_birth = None
            if isinstance(date_birth, datetime) and (
                    (datetime.now() - date_birth).days + 5) / 365 >= self.min_age_person:
                result = date_birth.strftime('%d.%m.%Y'),
        return result


class SerNumPassport(BaseField):
    words_to_search_in_title = ['паспортные', 'данные']
    or_words_to_search_in_title = ['серия', 'номер', 'паспорта']
    max_years_after_date_issue = 30

    def get_data(self, cell_value: str) -> tuple:
        result = False,
        cell_value = cell_value.replace(' ', '')
        if cell_value.isdigit() and len(cell_value) == 10:
            result = cell_value,
        return result


class DateIssuePassport(BaseField):
    words_to_search_in_title = ['дата', 'выдачи', 'паспорта']
    or_words_to_search_in_title = ['дата', 'выдачи', 'паспорт']
    max_years_after_date_issue = 30

    def get_data(self, cell_value: str) -> tuple:
        result = False,
        if cell_value:
            try:
                date_issue = datetime.strptime(cell_value, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    date_issue = datetime.strptime(cell_value, '%d.%m.%Y')
                except ValueError:
                    date_issue = None
            if isinstance(date_issue, datetime) and (
                    datetime.now() - date_issue).days / 365 <= self.max_years_after_date_issue:
                result = date_issue.strftime('%d.%m.%Y'),
        return result


class Checker:
    trans_fields = {
        'fullname': _('Fullname'),
        'date_birth': _('Date birth'),
        'date_issue_pass': _('Date issue passport'),
        'ser_num_pass': _('Serial and passport number'),
    }
    rows_for_check = 30
    max_title_row = 10

    def __init__(self):
        self.fields_table = {
            'fullname': {'column': None, 'row': None, 'class_field': FullNamePerson()},
            'date_birth': {'column': None, 'row': None, 'class_field': DateBirthPerson()},
            'date_issue_pass': {'column': None, 'row': None, 'class_field': DateIssuePassport()},
            'ser_num_pass': {'column': None, 'row': None, 'class_field': SerNumPassport()},
        }
        self.detected_columns = set()

    def check_field(self, field, title_row, column, sheet) -> bool:
        class_field = self.fields_table[field]['class_field']
        cell_value = str(sheet.cell(row=title_row, column=column).value).strip()
        if result_column := class_field.check_title(cell_value=cell_value, column=column):
            for data_row in range(
                    self.rows_for_check if sheet.max_row >= self.rows_for_check else sheet.max_row, 1, -1):
                cell_value = str(sheet.cell(row=data_row, column=column).value).strip()
                if result_row := class_field.check_data(cell_value=cell_value, row=data_row):
                    self.fields_table[field].update(result_column | result_row)
        if result := all((self.fields_table[field]['column'], self.fields_table[field]['row'])):
            self.detected_columns.add(result_column['column'])
        return result

    @classmethod
    def search_title_gen(cls, max_column: int) -> Generator:
        for title_row in range(1, cls.max_title_row):
            for column in range(1, max_column):
                yield title_row, column

    def check_fields_result(self, sheet):
        bad_result = _('not found')
        column = _('column')

        fc = self.fields_table["fullname"]["column"]
        fc_str = f'{fc if fc else ""} [ {sheet.cell(row=1, column=fc).column_letter if fc else bad_result} ]'

        dbc = self.fields_table["date_birth"]["column"]
        dbc_str = f'{dbc if dbc else ""} [ {sheet.cell(row=1, column=dbc).column_letter if dbc else bad_result} ]'

        dipc = self.fields_table["date_issue_pass"]["column"]
        dipc_str = f'{dipc if dipc else ""} [ {sheet.cell(row=1, column=dipc).column_letter if dipc else bad_result} ]'

        snpc = self.fields_table.get("ser_num_pass")["column"]
        snpc_str = f'{snpc if snpc else ""} [ {sheet.cell(row=1, column=snpc).column_letter if snpc else bad_result} ]'

        result = f'<ul><li>{self.trans_fields["fullname"]} - {column}: {fc_str}</li>' \
                 f'<li>{self.trans_fields["date_birth"]} - {column}: {dbc_str}</li>' \
                 f'<li>{self.trans_fields["date_issue_pass"]} - {column}: {dipc_str}</li>' \
                 f'<li>{self.trans_fields["ser_num_pass"]} - {column}: {snpc_str}</li></ul>'

        if all((fc, dbc, dipc, dipc_str)):
            return result
        else:
            raise FileVerificationException(message=result)


@shared_task(bind=True)
def check_fields(self, path, language=None):
    prev_language = translation.get_language()
    language and translation.activate(language)

    checker = Checker()
    progress_recorder = ProgressRecorder(self)

    sheet: Worksheet = openpyxl.load_workbook(path).active
    for num, (field, data) in enumerate(checker.fields_table.items(), 1):
        progress_recorder.set_progress(
            num,
            len(checker.fields_table.keys()),
            description=_('Searching') + f': {checker.trans_fields[field]}'
        )
        if not data['column'] and not data['row']:
            for title_row, column in checker.search_title_gen(max_column=sheet.max_column):
                if column not in checker.detected_columns:
                    if checker.check_field(field=field, title_row=title_row, column=column, sheet=sheet):
                        break

    result = checker.check_fields_result(sheet)
    translation.activate(prev_language)
    return result


def load_data_file(request: HttpRequest) -> tuple[str, AsyncResult]:
    date = datetime.strftime(datetime.now(), '%d.%m.%Y-%H:%M:%S')
    task = None

    if request.FILES:
        file: TemporaryUploadedFile | InMemoryUploadedFile = request.FILES['datafile']

        if file.name.endswith(('.xls', '.xlsx')):
            file_system = FileSystemStorage()
            filename = file_system.save(
                f'{request.path.lstrip("/")}date-{date}_user-pk-{request.user.pk}_filename-{file.name}', file)
            task: AsyncResult = check_fields.delay(path=f'media/{filename}', language=translation.get_language())
            redis_cache.set(get_redis_key(request=request, key_type='file_verification'), task.task_id)
            msg = _(f"File verification") + f": {file.name}"
        else:
            msg = _('Unsupported file, .xls .xlsx only')
    else:
        msg = _("Select File")
    return msg, task


